# -*- coding: utf-8 -*-
"""
Created on Fri Sep  9 17:22:13 2022
  l
@author: klecznar
"""
import sys

print("")
print("        __|__     ")
print("   ---o--(_)--o---")
print("          _   ____     ___    ___    ____    ____")
print("         | | |  _ \   / __|  / _ \  |  __|  / _  |")
print("         | | | | | | | (__  | (_) | | |    | (_| |")
print("         |_| |_| |_|  \___|  \___/  |_|    \___,_|")

print("\n       WELCOME TO INCORA'S WHITELIST WEB SCRAPER!")
print(" ")

# used to interact with excel file
import openpyxl
import pandas as pd
from functions import locate_col

# used to interact with website
import requests
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.wait import WebDriverWait

# used to interact with downloaded files and xlsx file
import datetime
import time
import os
import glob
import pathlib
from pathlib import Path
import shutil

# GET PATH TO USER'S DOWNLOADS DIRECTORY
user_path = str(Path.home() / "Downloads")

#OPENING EXCEL FILE

print()
path = input("Please provide full path to the Excel file: ")
print("Loading Excel file ...")

try:
    vendor_list = openpyxl.load_workbook(path)
    print("\n Success! Excel file is loaded.")
    print("\nNow let me look through the file...")
except Exception:
    print("Cannot open Excel file, please make sure to input full path to the Excel file...")
    time.sleep(10)


# Get list of items to search:
# NIP,
# bank account,

sheet = vendor_list.sheetnames
NIP_col = locate_col(vendor_list, sheet[0], 'VTID05')
bank_account_col = locate_col(vendor_list, sheet[0], 'BKAC05')  # Used for web search
supplier_name_col = locate_col(vendor_list, sheet[0], 'SNAM05')
supplier_code_col = locate_col(vendor_list, sheet[0], 'SUPN05')

file_name = "Vendor List.xlsx"
xl_workbook = pd.ExcelFile(file_name)  # Load the excel workbook
df = xl_workbook.parse("Sheet1")  # Parse the sheet into a dataframe

NIP_list = [] # Cast the desired column into a python list
bank_account_list = []

#lists to hold my results
supplier_name_list_nip = []
supplier_code_list_nip = []
supplier_name_list_bank = []
supplier_code_list_bank = []


for i in range(1048575):  # 1,048,576 is max number of rows in excel
    NIP_col_item_ID = vendor_list['Sheet1'].cell(row=2+i, column=NIP_col).value
    supplier_name_item1 = vendor_list['Sheet1'].cell(row=2+i, column=supplier_name_col).value
    supplier_code_item1 = vendor_list['Sheet1'].cell(row=2+i, column=supplier_code_col).value
    if NIP_col_item_ID is not None and NIP_col_item_ID != '':
        if len(NIP_col_item_ID) > 5:
            NIP_list.append(NIP_col_item_ID.replace(" ",""))
            supplier_name_list_nip.append(supplier_name_item1)
            supplier_code_list_nip.append(supplier_code_item1)


for y in range(1048575):
    bank_account_item_ID = vendor_list['Sheet1'].cell(row=2+y, column=bank_account_col).value
    supplier_name_item2 = vendor_list['Sheet1'].cell(row=2+y, column=supplier_name_col).value
    supplier_code_item2 = vendor_list['Sheet1'].cell(row=2+y, column=supplier_code_col).value
    if bank_account_item_ID is not None and bank_account_item_ID != '':
        if len(bank_account_item_ID) > 5:
            if not str(bank_account_item_ID).startswith('00000'): 
                bank_account_list.append(bank_account_item_ID.replace(" ",""))
                supplier_name_list_bank.append(supplier_name_item2)
                supplier_code_list_bank.append(supplier_code_item2)

#  CREATE DESTINATION FOLDER TO KEEP REPORTS
cd = datetime.datetime.now().strftime('%Y-%m-%d %H;%M;%S')  # get current datetime
parent_dir = (r'C:\Users\karol\OneDrive\Pulpit\whitelist\potwierdzenie')
# parent_dir = (r'\\components\data\GBD-Shares\GBDFI\9.2 Poland Purchase Ledger\9.1.0 Overhead Purchase Orders\white list\Potwierdzenia')
dest_dir = os.path.join(parent_dir, cd)
dd = os.mkdir(dest_dir)

# Calculate workload & print summary                   
NIP_searches = len(NIP_list)
print("Total NIP numbers: ", NIP_searches)
bank_account_searches = len(bank_account_list) 
print("Total bank account numbers: ", bank_account_searches)

total_searches = NIP_searches + bank_account_searches
manual_processing_time = 20  # Seconds - time saving
processing_time = round((total_searches * 10) / 60) # minutes

print("\n\n   SUMMARY:")
print("\n     NUMBER OF SEARCHES TO COMPLETE:",str(total_searches))
print("\n     POTENTIAL PROCESSING TIME:",str(processing_time),"MINUTES")
print("\n\n   STARTING WEB SEARCH... PLEASE BE PATIENT...\n\n")

# used to avoid server problems, such as CORS policy
options = webdriver.ChromeOptions()
options.add_argument("--disable-web-security")
options.add_argument("--disable-gpu")
options.add_argument('--log-level=1')
options.add_argument("--ignore-certificate-errors")

#OPEN BROWSER
try:
    URL = 'https://www.podatki.gov.pl/wykaz-podatnikow-vat-wyszukiwarka'
    driver = webdriver.Chrome(options=options)
    driver.get(URL)
    page = requests.get(URL)
    soup = BeautifulSoup(page.content, 'html.parser')
except Exception:
    print('\n   ERROR: WEB BROWSER NOT COMPATIBLE WITH DRIVER\n')
    print('\n   TRY AGAIN AFTER REPLACING THE DRIVER FILE...\n')
    time.sleep(10)


#locating search buttons
elementTwo = driver.find_element(By.XPATH, '//button[@aria-label="Szukaj"][@id="sendTwo"]')
elementOne = driver.find_element(By.XPATH, '//button[@aria-label="Szukaj"][@id="sendOne"]')

# create lists to hold my results
NIP_list_results = []
bank_account_list_results = []

# LOOP THROUGH supplier NIP numbers & BANK ACCOUNTS

for item in NIP_list:
    WebDriverWait(driver, 2).until(EC.element_to_be_clickable((
        By.XPATH, '//label[@aria-label="Wpisz numer nip"]'
    ))).click()
    WebDriverWait(driver, 2).until(EC.element_to_be_clickable((
        By.XPATH, '//input[@placeholder="Wpisz numer nip"]'
    ))).click()
    WebDriverWait(driver, 2).until(EC.element_to_be_clickable((
        By.XPATH, '//input[@placeholder="Wpisz numer nip"]'
    ))).send_keys(item)
    time.sleep(2)
    try:
        if elementTwo.is_enabled():
            WebDriverWait(driver, 2).until(EC.element_to_be_clickable((
                By.XPATH, '//button[@aria-label="Szukaj"][@id="sendTwo"]'
            ))).click()
            figuruje_alert = WebDriverWait(driver, 5).until(EC.presence_of_element_located((
                By.XPATH, '//*[@id="tableOne"]/div[1]/div/h4'
            )))
            NIP_list_results.append(figuruje_alert.text)
            if figuruje_alert.text == "Figuruje w rejestrze VAT":
                drukuj_potwierdzenie_button = WebDriverWait(driver, 5).until(
                    EC.element_to_be_clickable((By.CSS_SELECTOR, '#superPrintButton'
                    ))).click()
                time.sleep(10)
                # RENAME DOWNLOADED FILE
                name_element = driver.find_element(By.CSS_SELECTOR, "#akmf-name > tbody > tr > td.twosecond")
                path = (user_path + '/*.pdf')  # * means all if need specific format then *.pdf
                list_of_files = glob.glob(path)
                latest_file = max(list_of_files, key=os.path.getctime)
                old_file_path = os.path.abspath(latest_file)
                new_name = name_element.text
                new_name_path = (user_path + '\\' + new_name + '.pdf')
                os.rename(old_file_path, new_name_path)
                # MOVE DOWNLOADED FILE
                new_folder_path = pathlib.Path(dest_dir)
                shutil.move(new_name_path, new_folder_path)
        elif elementTwo.is_enabled() is False:
            NIP_list_results.append("Nie znaleziono, podano błędny numer NIP")
    except:
        try:
            if WebDriverWait(driver, 2).until(EC.element_to_be_clickable((
                    By.XPATH, '//button[@aria-label="Szukaj"][@id="sendOne"]'
            ))).is_enabled():
                WebDriverWait(driver, 2).until(EC.element_to_be_clickable((
                    By.XPATH, '//button[@aria-label="Szukaj"][@id="sendOne"]'
                ))).click()
                figuruje_alert = WebDriverWait(driver, 5).until(EC.presence_of_element_located((
                    By.XPATH, '//*[@id="tableOne"]/div[1]/div/h4'
                )))
                #print(figuruje_alert.text)   #<-- sanity check
                NIP_list_results.append(figuruje_alert.text)
                if figuruje_alert.text == "Figuruje w rejestrze VAT":
                    drukuj_potwierdzenie_button = WebDriverWait(driver, 5).until(EC.element_to_be_clickable((
                        By.CSS_SELECTOR, '#superPrintButton'
                    ))).click()
                    time.sleep(10)
                    # RENAME DOWNLOADED FILE
                    name_element = driver.find_element(By.CSS_SELECTOR, "#akmf-name > tbody > tr > td.twosecond")
                    path = (user_path + '/*.pdf')  # * means all if need specific format then *.pdf
                    list_of_files = glob.glob(path)
                    latest_file = max(list_of_files, key=os.path.getctime)
                    old_file_path = os.path.abspath(latest_file)
                    new_name = name_element.text
                    new_name_path = (user_path + '\\' + new_name + '.pdf')
                    os.rename(old_file_path, new_name_path)
                    # MOVE DOWNLOADED FILE
                    new_folder_path = pathlib.Path(dest_dir)
                    shutil.move(new_name_path, new_folder_path)
                    print_popup = WebDriverWait(driver, 5).until(EC.element_to_be_clickable((
                        By.CLASS_NAME, 'action-button'
                    )))
                    if print_popup.is_enabled():
                        print_popup.click()
        except:
            if WebDriverWait(driver, 5).until(EC.presence_of_element_located((
                    By.XPATH, '//*[@id="errorBox"]/div/div[1]/h4'
            ))).text == "Nieprawidłowy NIP.":
                NIP_list_results.append(WebDriverWait(driver, 5).until(EC.presence_of_element_located((
                    By.XPATH, '//*[@id="errorBox"]/div/div[1]/h4'
                ))).text)
            if WebDriverWait(driver, 5).until(EC.presence_of_element_located((
                    By.XPATH, '//*[@id="errorBox"]/div/div[1]/h4'
            ))).text == "Nie figuruje w rejestrze VAT":
                NIP_list_results.append(WebDriverWait(driver, 5).until(EC.presence_of_element_located((
                    By.XPATH, '//*[@id="errorBox"]/div/div[1]/h4'
                ))).text)

print("NIP search completed!")


for account in bank_account_list:
    WebDriverWait(driver, 2).until(EC.element_to_be_clickable((
        By.XPATH, '//label[@aria-label="Numer konta"]'
    ))).click()
    WebDriverWait(driver, 2).until(EC.element_to_be_clickable((
        By.XPATH, '//input[@placeholder="Wpisz numer konta"]'
    ))).clear()
    WebDriverWait(driver, 2).until(EC.element_to_be_clickable((
        By.XPATH, '//input[@placeholder="Wpisz numer konta"]'
    ))).click()
    WebDriverWait(driver, 2).until(EC.element_to_be_clickable((
        By.XPATH, '//input[@placeholder="Wpisz numer konta"]'
    ))).send_keys(account)
    time.sleep(2)
    try:
        if elementTwo.is_enabled():
            WebDriverWait(driver, 2).until(EC.element_to_be_clickable((
                By.XPATH, '//button[@aria-label="Szukaj"][@id="sendTwo"]'
            ))).click()
            figuruje_alert = WebDriverWait(driver, 5).until(EC.presence_of_element_located((
                By.XPATH, '//*[@id="tableOne"]/div[1]/div/h4'
            )))
            bank_account_list_results.append(figuruje_alert.text)
            if figuruje_alert.text == "Figuruje w rejestrze VAT":
                drukuj_potwierdzenie_button = WebDriverWait(driver, 5).until(
                    EC.element_to_be_clickable((By.CSS_SELECTOR, '#superPrintButton'
                                                ))).click()
                # RENAME DOWNLOADED FILE
                name_element = driver.find_element(By.CSS_SELECTOR, "#akmf-name > tbody > tr > td.twosecond")
                path = (user_path + '/*.pdf')  # * means all if need specific format then *.pdf
                list_of_files = glob.glob(path)
                latest_file = max(list_of_files, key=os.path.getctime)
                old_file_path = os.path.abspath(latest_file)
                new_name = name_element.text
                new_name_path = (user_path + '\\' + new_name + '.pdf')
                os.rename(old_file_path, new_name_path)
                # MOVE DOWNLOADED FILE
                new_folder_path = pathlib.Path(dest_dir)
                shutil.move(new_name_path, new_folder_path)
        elif elementTwo.is_enabled() is False:
            bank_account_list_results.append("Nie znaleziono, podano błędny numer konta bankowego")
    except:
        try:
            if WebDriverWait(driver, 2).until(EC.element_to_be_clickable((
                    By.XPATH, '//button[@aria-label="Szukaj"][@id="sendOne"]'
            ))).is_enabled():
                WebDriverWait(driver, 2).until(EC.element_to_be_clickable((
                    By.XPATH, '//button[@aria-label="Szukaj"][@id="sendOne"]'
                ))).click()
                figuruje_alert = WebDriverWait(driver, 5).until(EC.presence_of_element_located((
                    By.XPATH, '//*[@id="tableOne"]/div[1]/div/h4'
                )))
                bank_account_list_results.append(figuruje_alert.text)
                if figuruje_alert.text == "Figuruje w rejestrze VAT":
                    drukuj_potwierdzenie_button = WebDriverWait(driver, 5).until(EC.element_to_be_clickable((
                        By.CSS_SELECTOR, '#superPrintButton'
                    ))).click()
                    # RENAME DOWNLOADED FILE
                    name_element = driver.find_element(By.CSS_SELECTOR, "#akmf-name > tbody > tr > td.twosecond")
                    path = (user_path + '/*.pdf')  # * means all if need specific format then *.pdf
                    list_of_files = glob.glob(path)
                    latest_file = max(list_of_files, key=os.path.getctime)
                    old_file_path = os.path.abspath(latest_file)
                    new_name = name_element.text
                    new_name_path = (user_path + '\\' + new_name + '.pdf')
                    os.rename(old_file_path, new_name_path)
                    # MOVE DOWNLOADED FILE
                    new_folder_path = pathlib.Path(dest_dir)
                    shutil.move(new_name_path, new_folder_path)
                    print_popup = WebDriverWait(driver, 5).until(EC.element_to_be_clickable((
                        By.CLASS_NAME, 'action-button'
                    )))
                    if print_popup.is_enabled():
                        print_popup.click()
        except:
            if WebDriverWait(driver, 5).until(EC.presence_of_element_located((
                    By.XPATH, '//*[@id="errorBox"]/div/div[1]/h4'
            ))).text == "Nieprawidłowy numer konta bankowego.":
                bank_account_list_results.append(WebDriverWait(driver, 5).until(EC.presence_of_element_located((
                    By.XPATH, '//*[@id="errorBox"]/div/div[1]/h4'
                ))).text)
            if WebDriverWait(driver, 5).until(EC.presence_of_element_located((
                    By.XPATH, '//*[@id="errorBox"]/div/div[1]/h4'
            ))).text == "Rachunek nie figuruje na wykazie":
                bank_account_list_results.append(WebDriverWait(driver, 5).until(EC.presence_of_element_located((
                    By.XPATH, '//*[@id="errorBox"]/div/div[1]/h4'
                ))).text)

driver.quit()  # close website
print("Bank account search completed!")

# end web scraping
# save data to excel
df_bank = pd.DataFrame(
    {
        'Supplier Code': supplier_code_list_bank,
        'Supplier name': supplier_name_list_bank,
        'Bank Account': bank_account_list,
        'is_registered?':bank_account_list_results
    }
)
df_nip = pd.DataFrame(
    {
        'Supplier Code': supplier_code_list_nip,
        'Supplier name': supplier_name_list_nip,
        'NIP': NIP_list,
        'is_registered?': NIP_list_results
    }
) # create DFs with the lists

# CREATE XLSX REPORT --> one file, divided by sheets

date_string = datetime.datetime.now().strftime('%Y%m%d %H;%M;%S')
with pd.ExcelWriter("Scrapped Results_" + date_string +".xlsx") as writer:
    df_nip.to_excel(writer, sheet_name="NIP", index=False)
    df_bank.to_excel(writer, sheet_name="Bank Account", index=False)

# MOVE REPORT TO DESTINATION FOLDER

report_path = os.path.abspath(os.getcwd())

for file in os.listdir(report_path):
    if file.startswith('Scrapped Results_') and file.endswith('.xlsx'):
        shutil.move(file, new_folder_path)


print("\n All files downloaded successfuly!")

time.sleep(10)  # freeze time for 10 secs to enable reading above info